#!/usr/bin/env python3
"""
VMST UNEMPLOYMENT INGEST — Talnagögn_atvinnuleysi.xlsm
Downloads, verifies, and parses the VMST monthly unemployment dashboard into
long-form CSVs. All 5 data sheets (G1–G5) are extracted.

Source:
    https://island.is/s/vinnumalastofnun/maelabord-og-toelulegar-upplysingar
    → "Helstu talnagögn um atvinnuleysi" (XLSM on Contentful CDN)

Sheets:
    G1 — Árlegt atvinnuleysi á Íslandi (2000–2025)
    G2 — Mánaðarlegt atvinnuleysi, % og meðalfjöldi (2000–2026)
    G3 — Fjöldi á skrá eftir landsvæðum og sveitarfélögum (2000–2026)
    G4 — Fjöldi á skrá eftir aldri og lengd á skrá (2000–2026)
    G5 — Atvinnuleysi eftir atvinnugreinum (2000–2026)

Provenance trail (FOIA reply #05):
    VMST (Ásta Ásgeirsdóttir) — 20. apríl 2026 — vísaði á mælaborðið.
    Request: 02_VMST_SECTOR_UNEMPLOYMENT, sent 14. apríl 2026.

Usage:
    python vmst_ingest.py --download     # Fetch XLSM + verify checksum
    python vmst_ingest.py --parse        # Parse all sheets → CSV
    python vmst_ingest.py --all          # Download + parse + manifest
    python vmst_ingest.py --manifest     # (Re)compute SHA-256 manifest

Output: CSV files in ../data/raw/
    vmst_G1_unemployment_annual.csv
    vmst_G2_unemployment_monthly_rate.csv
    vmst_G3_register_by_region.csv
    vmst_G4_register_by_age.csv
    vmst_G5_register_by_sector.csv
    vmst_MANIFEST.md
"""

import argparse
import csv
import hashlib
import sys
from datetime import datetime, timezone
from pathlib import Path
from urllib.request import Request, urlopen
from urllib.error import URLError

SOURCE_URL = (
    "https://assets.ctfassets.net/8k0h54kbe6bj/"
    "688FRtXuoA4qkPXerKcuMT/e6b9794d1bdfcf5cb5747a46b9b3d836/"
    "Talnagogn_atvinnuleysi.xlsm"
)
DASHBOARD_URL = "https://island.is/s/vinnumalastofnun/maelabord-og-toelulegar-upplysingar"
POWERBI_URL = (
    "https://app.powerbi.com/view?r=eyJrIjoiZTc0NTIxYmItZTUwMS00YjAyLThhYTIt"
    "MDhhOGJiODRkMDg3IiwidCI6Ijc2NGEzMDZkLTBhNjgtNDVhZC05ZjA3LTZmMTgwNDQ0N2Nk"
    "NCIsImMiOjh9&pageName=ReportSection7e7dca64570c18a74eb9"
)

ROOT = Path(__file__).parent.parent
RAW_DIR = ROOT / "data" / "raw"
XLSM_PATH = RAW_DIR / "vmst_Talnagogn_atvinnuleysi_2026-03.xlsm"


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def download(target: Path, force: bool = False) -> str:
    if target.exists() and not force:
        print(f"  [skip] already present: {target.name}")
        return sha256(target)
    print(f"  [fetch] {SOURCE_URL}")
    req = Request(SOURCE_URL, headers={"User-Agent": "atvinnustefna-2035/1.0"})
    with urlopen(req, timeout=60) as resp:
        data = resp.read()
    target.write_bytes(data)
    digest = sha256(target)
    print(f"  [save ] {target} ({len(data):,} bytes)")
    print(f"  [sha  ] {digest}")
    return digest


def _dates(row: tuple, start_col: int) -> list[str | None]:
    out: list[str | None] = []
    for c in row[start_col:]:
        if isinstance(c, datetime):
            out.append(c.strftime("%Y-%m"))
        else:
            out.append(None)
    return out


def parse_g1(wb, out: Path) -> int:
    """G1: Árlegt atvinnuleysi. Columns: ar, atvinnuleysi_pct"""
    ws = wb["G1"]
    rows = list(ws.iter_rows(values_only=True))
    # Row 4 = header "Ár" | "Atvinnuleysi"; data from row 5 onward
    with open(out, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["ar", "atvinnuleysi_hlutfall"])
        n = 0
        for r in rows[5:]:
            if r[0] is None:
                continue
            year, rate = r[0], r[1]
            if isinstance(year, (int, float)) and rate is not None:
                w.writerow([int(year), rate])
                n += 1
    return n


def parse_g2(wb, out: Path) -> int:
    """G2: Mánaðarlegt atvinnuleysi %. First col = flokkur, dates from col 2."""
    ws = wb["G2"]
    rows = list(ws.iter_rows(values_only=True))
    dates = _dates(rows[6], 2)
    # Two sections: "Atvinnuleysi %" (row 7) and "Atvinnulausir, meðalfjöldi" (later)
    current_metric: str | None = None
    with open(out, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["manudur", "maeling", "flokkur", "gildi"])
        n = 0
        for r in rows[7:]:
            label = r[0]
            if label is None:
                continue
            label_s = str(label).strip()
            if label_s in (
                "Atvinnuleysi %",
                "Atvinnulausir, meðalfjöldi á mánuði",
            ):
                current_metric = label_s
                continue
            if current_metric is None:
                continue
            for j, val in enumerate(r[2:]):
                if j >= len(dates) or dates[j] is None:
                    continue
                if val is None or val == "" or val == "..":
                    continue
                w.writerow([dates[j], current_metric, label_s, val])
                n += 1
    return n


def parse_g3(wb, out: Path) -> int:
    """G3: Fjöldi á skrá eftir landsvæðum og sveitarfélögum.
    Col A = svæði/landshluti, Col B = kyn, dates from col 3."""
    ws = wb["G3"]
    rows = list(ws.iter_rows(values_only=True))
    dates = _dates(rows[5], 3)
    with open(out, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["manudur", "svaedi", "kyn_eda_rikisfang", "fjoldi"])
        n = 0
        current_area: str | None = None
        for r in rows[6:]:
            col_a, col_b = r[0], r[1]
            if col_a is not None and str(col_a).strip() not in (
                "Atvinnulausir, fjöldi í lok mánaðar",
            ):
                current_area = str(col_a).strip()
            sub = None if col_b is None else str(col_b).strip()
            if current_area is None or current_area == "Atvinnulausir, fjöldi í lok mánaðar":
                continue
            for j, val in enumerate(r[3:]):
                if j >= len(dates) or dates[j] is None:
                    continue
                if val is None or val == "" or val == "..":
                    continue
                w.writerow([dates[j], current_area, sub or "", val])
                n += 1
    return n


def parse_g4(wb, out: Path) -> int:
    """G4: Fjöldi eftir aldri og lengd á skrá.
    Col A = rikisfang/heild, Col B = flokkur (aldur/lengd)."""
    ws = wb["G4"]
    rows = list(ws.iter_rows(values_only=True))
    dates = _dates(rows[5], 3)
    current_citz: str | None = None
    with open(out, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["manudur", "rikisfang", "flokkur", "fjoldi"])
        n = 0
        for r in rows[6:]:
            col_a, col_b = r[0], r[1]
            a = None if col_a is None else str(col_a).strip()
            b = None if col_b is None else str(col_b).strip()
            if a and a != "Atvinnulausir eftir aldri":
                current_citz = a
                # if this row also has col_b it is a data row for the group header
            if current_citz is None or a == "Atvinnulausir eftir aldri":
                continue
            flokkur = b
            if flokkur is None:
                continue
            for j, val in enumerate(r[3:]):
                if j >= len(dates) or dates[j] is None:
                    continue
                if val is None or val == "" or val == "..":
                    continue
                w.writerow([dates[j], current_citz, flokkur, val])
                n += 1
    return n


def parse_g5(wb, out: Path) -> int:
    """G5: Atvinnuleysi eftir atvinnugreinum.
    Three blocks: Allir (r7-25), Íslenskt (r26-44), Erlent (r45-63)."""
    ws = wb["G5"]
    rows = list(ws.iter_rows(values_only=True))
    dates = _dates(rows[5], 2)
    sections = {
        "Allir": (7, 25),
        "Íslenskt ríkisfang": (26, 44),
        "Erlent ríkisfang": (45, 63),
    }
    with open(out, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["manudur", "rikisfang", "atvinnugrein", "fjoldi"])
        n = 0
        for citz, (start, end) in sections.items():
            for r in rows[start:end + 1]:
                sector = r[0]
                if sector is None:
                    continue
                sector_s = str(sector).strip()
                if sector_s == citz:
                    continue  # block header
                for j, val in enumerate(r[2:]):
                    if j >= len(dates) or dates[j] is None:
                        continue
                    if val is None or val == "" or val == "..":
                        continue
                    w.writerow([dates[j], citz, sector_s, val])
                    n += 1
    return n


def parse_all() -> dict[str, tuple[Path, int]]:
    try:
        import openpyxl  # type: ignore
    except ImportError:
        print("ERROR: openpyxl required. pip install openpyxl", file=sys.stderr)
        sys.exit(1)
    print(f"  [open ] {XLSM_PATH}")
    wb = openpyxl.load_workbook(XLSM_PATH, data_only=True, keep_vba=False)
    outputs = {
        "G1": (RAW_DIR / "vmst_G1_unemployment_annual.csv", parse_g1),
        "G2": (RAW_DIR / "vmst_G2_unemployment_monthly_rate.csv", parse_g2),
        "G3": (RAW_DIR / "vmst_G3_register_by_region.csv", parse_g3),
        "G4": (RAW_DIR / "vmst_G4_register_by_age.csv", parse_g4),
        "G5": (RAW_DIR / "vmst_G5_register_by_sector.csv", parse_g5),
    }
    result: dict[str, tuple[Path, int]] = {}
    for sheet, (path, fn) in outputs.items():
        n = fn(wb, path)
        print(f"  [parse] {sheet} → {path.name}  ({n:,} rows)")
        result[sheet] = (path, n)
    return result


def write_manifest(xlsm_sha: str, parsed: dict[str, tuple[Path, int]]) -> Path:
    path = RAW_DIR / "vmst_MANIFEST.md"
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    lines = [
        "# VMST Unemployment Data — Manifest",
        "",
        f"**Generated:** {now}",
        f"**Ingest script:** `scripts/vmst_ingest.py`",
        "",
        "## Source",
        "",
        f"- Dashboard: <{DASHBOARD_URL}>",
        f"- Power BI (G5 page): <{POWERBI_URL}>",
        f"- XLSM file: <{SOURCE_URL}>",
        "",
        "## Provenance",
        "",
        "FOIA reply #05 — VMST (Ásta Ásgeirsdóttir, deildarstjóri Greiningardeildar) 20. apríl 2026.",
        "Request 02_VMST_SECTOR_UNEMPLOYMENT (sent 14. apríl 2026) was answered by",
        "referring to the public dashboard (self-service — no bespoke query).",
        "",
        "## Files",
        "",
        f"| File | Rows | SHA-256 |",
        f"|------|-----:|---------|",
        f"| `{XLSM_PATH.name}` | — | `{xlsm_sha}` |",
    ]
    for sheet, (p, n) in parsed.items():
        lines.append(f"| `{p.name}` | {n:,} | `{sha256(p)}` |")
    lines += [
        "",
        "## Sheet descriptions",
        "",
        "| Sheet | Content | Coverage | Columns |",
        "|-------|---------|----------|---------|",
        "| G1 | Árlegt atvinnuleysi (%) | 2000–2025 | ar, atvinnuleysi_hlutfall |",
        "| G2 | Mánaðarlegt atvinnuleysi (% og meðalfjöldi) eftir landshluta, kyni, ríkisfangi | 2000-02 → 2026-03 | manudur, maeling, flokkur, gildi |",
        "| G3 | Fjöldi á skrá eftir svæði/sveitarfélagi × kyni/ríkisfangi | 2000-02 → 2026-03 | manudur, svaedi, kyn_eda_rikisfang, fjoldi |",
        "| G4 | Fjöldi á skrá eftir aldurshópum og lengd á skrá × ríkisfangi | 2000-02 → 2026-03 | manudur, rikisfang, flokkur, fjoldi |",
        "| G5 | Fjöldi á skrá eftir **atvinnugreinum** × ríkisfangi | 2000-02 → 2026-03 | manudur, rikisfang, atvinnugrein, fjoldi |",
        "",
        "## Known limitations (G5 sector breakdown)",
        "",
        "- **Not ÍSAT-compliant.** VMST uses 18 bespoke categories. Primary correspondence:",
        "  - ÍSAT J = 'Upplýsingar og fjarskipti' (1:1) ✅",
        "  - ÍSAT K = 'Fjármála- og vátryggingastarfsemi' (1:1) ✅",
        "  - ÍSAT M = 'Sérfræðileg, vísindaleg og tæknileg starfsemi' (1:1) ✅",
        "  - ÍSAT F = 'Byggingastarfsemi og mannvirkjagerð' (1:1) ✅",
        "  - ÍSAT I = 'Gistiþjónusta' + 'Veitingaþjónusta' + 'Ferðaþjónusta ýmis' (3:1) ⚠️",
        "  - ÍSAT O–Q = 'Opinber þjónusta, fræðsla, heilbrigðis- og félagsþjónusta' (many:1) ⚠️",
        "  - **ÍSAT N (administrative & support services) has no mapping — absorbed in 'Ýmis þjónustustarfsemi'** ❌",
        "- **Head counts only** — no labour-force denominators by sector → no sector-level unemployment rate computable from G5 alone.",
        "- **VBA macros stripped** (`keep_vba=False`) — parsed values are static snapshots (as of file publication date).",
        "",
        "## Reproducibility",
        "",
        "```",
        "cd data/iceland-ai-governance-audit",
        "python scripts/vmst_ingest.py --all",
        "```",
        "",
        "Expected SHA-256 of `vmst_Talnagogn_atvinnuleysi_2026-03.xlsm`:",
        "",
        f"    {xlsm_sha}",
        "",
        "*If the upstream XLSM is re-published, this hash will change and the manifest must be regenerated. The dashboard does not version historical files.*",
        "",
    ]
    path.write_text("\n".join(lines), encoding="utf-8")
    print(f"  [write] {path}")
    return path


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--download", action="store_true", help="Fetch XLSM only")
    p.add_argument("--parse", action="store_true", help="Parse sheets to CSV")
    p.add_argument("--manifest", action="store_true", help="Write MANIFEST")
    p.add_argument("--all", action="store_true", help="Download + parse + manifest")
    p.add_argument("--force", action="store_true", help="Re-download even if present")
    args = p.parse_args()

    if not any((args.download, args.parse, args.manifest, args.all)):
        p.print_help()
        return 1

    RAW_DIR.mkdir(parents=True, exist_ok=True)

    xlsm_sha = ""
    parsed: dict[str, tuple[Path, int]] = {}

    if args.download or args.all:
        print("== DOWNLOAD ==")
        xlsm_sha = download(XLSM_PATH, force=args.force)

    if args.parse or args.all:
        print("== PARSE ==")
        parsed = parse_all()

    if args.manifest or args.all:
        print("== MANIFEST ==")
        if not xlsm_sha:
            xlsm_sha = sha256(XLSM_PATH)
        if not parsed:
            # compute row counts for existing CSVs
            names = {
                "G1": RAW_DIR / "vmst_G1_unemployment_annual.csv",
                "G2": RAW_DIR / "vmst_G2_unemployment_monthly_rate.csv",
                "G3": RAW_DIR / "vmst_G3_register_by_region.csv",
                "G4": RAW_DIR / "vmst_G4_register_by_age.csv",
                "G5": RAW_DIR / "vmst_G5_register_by_sector.csv",
            }
            for k, pth in names.items():
                if pth.exists():
                    with open(pth) as f:
                        n = sum(1 for _ in f) - 1
                    parsed[k] = (pth, n)
        write_manifest(xlsm_sha, parsed)

    print("\n== OK ==")
    return 0


if __name__ == "__main__":
    sys.exit(main())
